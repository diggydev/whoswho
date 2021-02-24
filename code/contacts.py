import datetime
import os
import yaml
from openpyxl import Workbook


def cleanse_data(data):
    cleansed = dict()
    for item in data.items():
        cleansed[item[0].upper()] = item[1]
    return cleansed


headers = set()
rows = list()
for contact_filename in os.listdir('../individuals/'):
    if contact_filename.endswith('.yaml'):
        with open(os.path.join('individuals', contact_filename), 'r') as contact_file:
            contact_data = cleanse_data(yaml.load(contact_file, Loader=yaml.loader.FullLoader))
            rows.append(contact_data)
            headers = headers.union(contact_data.keys())


wb = Workbook()
ws = wb.active
row_index = 1
col_index = 1
for header in headers:
    ws.cell(row=row_index, column=col_index, value=header)
    col_index += 1

for row in rows:
    col_index = 1
    row_index += 1
    for header in headers:
        value = row.get(header, "")
        ws.cell(row=row_index, column=col_index, value=value)
        col_index += 1

wb.save(f'../excel/contacts_{datetime.datetime.now().strftime(format="%Y-%m-%dT%H%MZ")}.xlsx')
